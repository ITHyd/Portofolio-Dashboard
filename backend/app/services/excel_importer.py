"""Reusable importer for Portfolio_Register_v4.xlsx.

Both seed.py and the /api/imports endpoint call into this. The importer is
idempotent: projects, weekly status rows, resources, and resource weeks are
upserted; risks/CSAT are inserted only when no equivalent row exists.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Union

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    Csat,
    Project,
    Resource,
    ResourceWeek,
    RiskIssue,
    User,
    WeeklyStatus,
)

SheetSource = Union[str, Path, BytesIO]

SHEET_PROJECTS = "🗂 Project Register"
SHEET_STATUS = "🚦 Weekly Status Update"
SHEET_RISKS = "⚠️ Risks & Issues"
SHEET_CSAT = "⭐ Client Satisfaction"
SHEET_RESOURCES = "👥 Resource & Utilisation"


def _val(v: Any) -> Any:
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _as_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _resolve_project_id(ref: Any, name: Any, ref_to_id: dict[str, int]) -> int | None:
    if isinstance(ref, str) and ref in ref_to_id:
        return ref_to_id[ref]
    if isinstance(name, str) and name in ref_to_id:
        return ref_to_id[name]
    return None


def import_portfolio_register(db: Session, source: SheetSource) -> dict[str, int]:
    """Run the import. Returns a dict of inserted/updated counts per entity."""
    wb = openpyxl.load_workbook(source, data_only=True)

    counts = {
        "projects_inserted": 0,
        "projects_updated": 0,
        "weekly_status_upserted": 0,
        "risks_issues_inserted": 0,
        "csat_inserted": 0,
        "resources_inserted": 0,
        "resource_weeks_upserted": 0,
    }

    pm_user = db.scalar(select(User).where(User.email == "pm@nxzen.com"))
    cp_user = db.scalar(select(User).where(User.email == "cp@nxzen.com"))

    ref_to_id: dict[str, int] = {}
    for p in db.scalars(select(Project)).all():
        if p.ref:
            ref_to_id[p.ref] = p.id
        ref_to_id[p.name] = p.id

    if SHEET_PROJECTS in wb.sheetnames:
        ws = wb[SHEET_PROJECTS]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not row[1]:
                continue
            ref = _val(row[0])
            name = _val(row[1])
            client = _val(row[2])
            if not name or not client:
                continue
            ref_clean = ref if isinstance(ref, str) and ref != "0" else None
            fields = {
                "name": name,
                "client": client,
                "sub_proposition": _val(row[3]),
                "phase": _val(row[4]),
                "status": _val(row[5]) or "Active",
                "start_date": _as_date(row[6]),
                "end_date_baseline": _as_date(row[7]),
                "end_date_forecast": _as_date(row[8]),
                "days_slippage": row[9] if isinstance(row[9], int) else None,
                "contract_value_gbp": row[10] if isinstance(row[10], (int, float)) else None,
                "margin_target_pct": row[11] if isinstance(row[11], (int, float)) else None,
                "pm_name": _val(row[12]),
                "cp_name": _val(row[13]),
                "project_code": _val(row[14]) if len(row) > 14 else None,
                "notes": _val(row[15]) if len(row) > 15 else None,
            }

            existing: Project | None = None
            if ref_clean:
                existing = db.scalar(select(Project).where(Project.ref == ref_clean))
            if not existing:
                existing = db.scalar(
                    select(Project).where(Project.name == name, Project.client == client)
                )

            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                if ref_clean and not existing.ref:
                    existing.ref = ref_clean
                db.flush()
                ref_to_id[existing.name] = existing.id
                if existing.ref:
                    ref_to_id[existing.ref] = existing.id
                counts["projects_updated"] += 1
            else:
                p = Project(
                    ref=ref_clean,
                    pm_user_id=pm_user.id if pm_user else None,
                    cp_user_id=cp_user.id if cp_user else None,
                    **fields,
                )
                db.add(p)
                db.flush()
                ref_to_id[p.name] = p.id
                if p.ref:
                    ref_to_id[p.ref] = p.id
                counts["projects_inserted"] += 1
        db.commit()

    if SHEET_STATUS in wb.sheetnames:
        ws = wb[SHEET_STATUS]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not row[1]:
                continue
            project_id = _resolve_project_id(row[0], row[1], ref_to_id)
            week_ending = _as_date(row[3])
            if not project_id or not week_ending:
                continue
            data = {
                "schedule_rag": _val(row[4]),
                "resource_rag": _val(row[5]),
                "scope_rag": _val(row[6]),
                "budget_rag": _val(row[7]),
                "overall_rag": _val(row[8]),
                "key_flag_comment": _val(row[9]),
                "next_milestone": _val(row[10]),
                "milestone_due": _as_date(row[11]),
                "milestone_status": _val(row[12]),
            }
            existing = db.scalar(
                select(WeeklyStatus).where(
                    WeeklyStatus.project_id == project_id,
                    WeeklyStatus.week_ending == week_ending,
                )
            )
            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
            else:
                db.add(
                    WeeklyStatus(
                        project_id=project_id,
                        week_ending=week_ending,
                        updated_by_user_id=pm_user.id if pm_user else None,
                        **data,
                    )
                )
            counts["weekly_status_upserted"] += 1
        db.commit()

    if SHEET_RISKS in wb.sheetnames:
        ws = wb[SHEET_RISKS]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not row[1]:
                continue
            project_id = _resolve_project_id(row[0], row[1], ref_to_id)
            if not project_id:
                continue
            type_ = _val(row[4])
            description = _val(row[6])
            if not type_ and not description:
                continue
            date_raised = _as_date(row[11])
            dup = db.scalar(
                select(RiskIssue).where(
                    RiskIssue.project_id == project_id,
                    RiskIssue.description == description,
                    RiskIssue.date_raised == date_raised,
                )
            )
            if dup:
                continue
            db.add(
                RiskIssue(
                    project_id=project_id,
                    type=type_,
                    rating=_val(row[5]),
                    description=description,
                    impact_if_unmitigated=_val(row[7]),
                    mitigation_action=_val(row[8]),
                    owner=_val(row[9]),
                    status=_val(row[10]) or "Open",
                    date_raised=date_raised,
                )
            )
            counts["risks_issues_inserted"] += 1
        db.commit()

    if SHEET_CSAT in wb.sheetnames:
        ws = wb[SHEET_CSAT]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not row[1]:
                continue
            project_id = _resolve_project_id(row[0], row[1], ref_to_id)
            if not project_id:
                continue
            date_collected = _as_date(row[3])
            score = row[4] if isinstance(row[4], (int, float)) else None
            if not score and not date_collected:
                continue
            dup = db.scalar(
                select(Csat).where(
                    Csat.project_id == project_id, Csat.date_collected == date_collected
                )
            )
            if dup:
                continue
            db.add(
                Csat(
                    project_id=project_id,
                    date_collected=date_collected,
                    score_1_5=int(score) if score else None,
                    comment=_val(row[5]),
                    collected_by_name=_val(row[6]),
                    next_collection_due=_as_date(row[7]),
                )
            )
            counts["csat_inserted"] += 1
        db.commit()

    if SHEET_RESOURCES in wb.sheetnames:
        ws = wb[SHEET_RESOURCES]
        # The Excel sheet's "week ending" is not in the data rows; the sheet header
        # implies "this week". Use the latest weekly_status week as the import target,
        # falling back to today's Friday.
        target_week = db.scalar(select(WeeklyStatus.week_ending).order_by(WeeklyStatus.week_ending.desc()).limit(1))
        if not target_week:
            today = date.today()
            offset = (4 - today.weekday()) % 7
            from datetime import timedelta as _td

            target_week = today + _td(days=offset)

        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not row[1] or not isinstance(row[0], str):
                continue
            code = _val(row[0])
            name = _val(row[1])
            if not code or not name:
                continue
            resource = db.scalar(select(Resource).where(Resource.code == code))
            if resource:
                resource.name = name
                resource.practice = _val(row[2])
                resource.region = _val(row[3])
                resource.contract_hours_per_week = (
                    row[4] if isinstance(row[4], (int, float)) else None
                )
            else:
                resource = Resource(
                    code=code,
                    name=name,
                    practice=_val(row[2]),
                    region=_val(row[3]),
                    contract_hours_per_week=row[4] if isinstance(row[4], (int, float)) else None,
                    active=True,
                )
                db.add(resource)
                db.flush()
                counts["resources_inserted"] += 1

            rw_data = {
                "leave_hrs": row[5] if isinstance(row[5], (int, float)) else None,
                "billable_hrs": row[6] if isinstance(row[6], (int, float)) else None,
                "non_billable_hrs": row[7] if isinstance(row[7], (int, float)) else None,
                "utilisation_pct": row[8] if isinstance(row[8], (int, float)) else None,
                "assigned_project_refs": _val(row[9]),
                "assignment_status": _val(row[10]),
                "notes": _val(row[11]) if len(row) > 11 else None,
            }
            rw = db.scalar(
                select(ResourceWeek).where(
                    ResourceWeek.resource_id == resource.id,
                    ResourceWeek.week_ending == target_week,
                )
            )
            if rw:
                for k, v in rw_data.items():
                    setattr(rw, k, v)
            else:
                db.add(ResourceWeek(resource_id=resource.id, week_ending=target_week, **rw_data))
            counts["resource_weeks_upserted"] += 1
        db.commit()

    return counts
